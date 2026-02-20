#!/usr/bin/env python3
"""
Generate impressive Finance Dashboard Excel files for portfolio screenshots.
Creates executive-level financial dashboards with charts, KPIs, and variance analysis.
"""
import os
from datetime import datetime, timedelta
import random

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.formatting.rule import ColorScaleRule
except ImportError:
    print("Installing openpyxl...")
    os.system("pip install openpyxl")
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.formatting.rule import ColorScaleRule

output_dir = r"C:\Users\sorat\Desktop\Coding\portfolio_my\temp_excel_outputs"
os.makedirs(output_dir, exist_ok=True)

def create_executive_summary_dashboard():
    """Create comprehensive executive summary with KPIs and charts."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Executive Summary"

    # Title
    ws['A1'] = "EXECUTIVE FINANCIAL DASHBOARD - Q1 2025"
    ws['A1'].font = Font(size=18, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws.merge_cells('A1:L1')
    ws.row_dimensions[1].height = 35

    # KPI Section
    kpis = [
        ("Total Revenue", "AED 127.5M", "+12.3%", "00B050"),
        ("Gross Profit", "AED 48.2M", "+8.7%", "00B050"),
        ("EBITDA", "AED 18.9M", "-3.2%", "C00000"),
        ("Net Profit", "AED 12.1M", "+5.4%", "00B050"),
        ("Operating Margin", "37.8%", "+2.1pp", "00B050"),
        ("Cash Position", "AED 23.4M", "+15.6%", "00B050"),
    ]

    col = 1
    for label, value, change, color in kpis:
        # Label
        cell = ws.cell(3, col, label)
        cell.font = Font(size=10, bold=True, color="666666")
        cell.alignment = Alignment(horizontal="center")

        # Value
        cell = ws.cell(4, col, value)
        cell.font = Font(size=16, bold=True)
        cell.alignment = Alignment(horizontal="center")

        # Change
        cell = ws.cell(5, col, change)
        cell.font = Font(size=12, bold=True, color=color)
        cell.alignment = Alignment(horizontal="center")

        col += 2

    # Monthly P&L Summary
    ws['A8'] = "MONTHLY PROFIT & LOSS SUMMARY (AED '000)"
    ws['A8'].font = Font(size=14, bold=True)
    ws.merge_cells('A8:M8')

    headers = ['Account', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
               'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec', 'YTD']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(9, col, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # P&L Lines
    pl_lines = [
        ('Revenue', [10500, 11200, 10800, 11500, 10900, 11300, 11800, 11200, 10600, 11400, 10800, 11200]),
        ('Cost of Sales', [6300, 6720, 6480, 6900, 6540, 6780, 7080, 6720, 6360, 6840, 6480, 6720]),
        ('Gross Profit', [4200, 4480, 4320, 4600, 4360, 4520, 4720, 4480, 4240, 4560, 4320, 4480]),
        ('Operating Expenses', [2800, 2900, 2850, 2950, 2880, 2920, 2980, 2900, 2820, 2940, 2860, 2900]),
        ('EBITDA', [1400, 1580, 1470, 1650, 1480, 1600, 1740, 1580, 1420, 1620, 1460, 1580]),
        ('Depreciation', [350, 350, 350, 350, 350, 350, 350, 350, 350, 350, 350, 350]),
        ('EBIT', [1050, 1230, 1120, 1300, 1130, 1250, 1390, 1230, 1070, 1270, 1110, 1230]),
        ('Interest', [180, 180, 180, 180, 180, 180, 180, 180, 180, 180, 180, 180]),
        ('Net Profit', [870, 1050, 940, 1120, 950, 1070, 1210, 1050, 890, 1090, 930, 1050]),
    ]

    row = 10
    for line_name, values in pl_lines:
        ws.cell(row, 1, line_name)
        ws.cell(row, 1).font = Font(bold=True if line_name in ['Revenue', 'Gross Profit', 'EBITDA', 'EBIT', 'Net Profit'] else False)

        ytd = 0
        for col, val in enumerate(values, 2):
            ws.cell(row, col, val)
            ws.cell(row, col).number_format = '#,##0'
            ytd += val

            # Highlight negative values
            if line_name in ['Cost of Sales', 'Operating Expenses', 'Depreciation', 'Interest'] or val < 0:
                ws.cell(row, col).font = Font(color="C00000")

        ws.cell(row, 14, ytd)
        ws.cell(row, 14).number_format = '#,##0'
        ws.cell(row, 14).font = Font(bold=True)

        # Bottom border for subtotals
        if line_name in ['Gross Profit', 'EBITDA', 'EBIT', 'Net Profit']:
            for col in range(1, 15):
                ws.cell(row, col).border = Border(bottom=Side(style='medium'))

        row += 1

    # Variance Analysis Section
    ws['A21'] = "VARIANCE ANALYSIS - BUDGET vs ACTUAL"
    ws['A21'].font = Font(size=14, bold=True)
    ws.merge_cells('A21:F21')

    var_headers = ['KPI', 'Budget', 'Actual', 'Variance', 'Variance %', 'Status']
    for col, header in enumerate(var_headers, 1):
        cell = ws.cell(22, col, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    var_data = [
        ('Revenue', 125000, 127500),
        ('Gross Profit', 47500, 48200),
        ('EBITDA', 19500, 18900),
        ('Net Profit', 11800, 12100),
        ('OpEx Ratio', 23.5, 22.8),
    ]

    row = 23
    for kpi, budget, actual in var_data:
        ws.cell(row, 1, kpi)
        ws.cell(row, 2, budget)
        ws.cell(row, 2).number_format = '#,##0' if kpi != 'OpEx Ratio' else '0.0"%"'

        ws.cell(row, 3, actual)
        ws.cell(row, 3).number_format = '#,##0' if kpi != 'OpEx Ratio' else '0.0"%"'

        variance = actual - budget
        ws.cell(row, 4, variance)
        ws.cell(row, 4).number_format = '#,##0' if kpi != 'OpEx Ratio' else '0.0"%"'
        ws.cell(row, 4).font = Font(color="00B050" if variance > 0 and kpi != 'OpEx Ratio' else "C00000" if variance < 0 else "000000")

        var_pct = (variance / budget) * 100
        ws.cell(row, 5, var_pct)
        ws.cell(row, 5).number_format = '0.0"%"'
        ws.cell(row, 5).font = Font(color="00B050" if var_pct > 0 and kpi != 'OpEx Ratio' else "C00000" if var_pct < 0 else "000000")

        status = "✓ On Track" if abs(var_pct) < 5 else "⚠ Review Required"
        ws.cell(row, 6, status)
        ws.cell(row, 6).fill = PatternFill(start_color="C6EFCE" if abs(var_pct) < 5 else "FFEB9C", fill_type="solid")

        row += 1

    # Auto-fit columns
    for col in range(1, 15):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12

    ws.column_dimensions['A'].width = 20

    wb.save(f"{output_dir}/Executive_Summary_Dashboard.xlsx")
    print("✓ Created Executive_Summary_Dashboard.xlsx")

def create_cash_flow_analysis():
    """Create detailed cash flow statement with waterfall analysis."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cash Flow Analysis"

    # Title
    ws['A1'] = "CASH FLOW STATEMENT & LIQUIDITY ANALYSIS"
    ws['A1'].font = Font(size=18, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws.merge_cells('A1:F1')
    ws.row_dimensions[1].height = 35

    # Cash Flow Statement
    ws['A3'] = "CONSOLIDATED CASH FLOW STATEMENT (AED '000)"
    ws['A3'].font = Font(size=14, bold=True)
    ws.merge_cells('A3:F3')

    headers = ['Category', 'Q1 2024', 'Q2 2024', 'Q3 2024', 'Q4 2024', 'Q1 2025']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(4, col, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    cf_items = [
        ('Operating Activities', []),
        ('Net Profit', [3200, 3450, 3100, 3600, 3850]),
        ('Depreciation & Amortization', [1050, 1050, 1050, 1050, 1050]),
        ('Change in Working Capital', [-850, -920, -780, -1100, -950]),
        ('Net Operating Cash Flow', [3400, 3580, 3370, 3550, 3950]),
        ('', []),
        ('Investing Activities', []),
        ('Capital Expenditures', [-2100, -1800, -2400, -2200, -1900]),
        ('Asset Disposals', [120, 0, 85, 150, 95]),
        ('Net Investing Cash Flow', [-1980, -1800, -2315, -2050, -1805]),
        ('', []),
        ('Financing Activities', []),
        ('Debt Drawdown/(Repayment)', [1500, -800, 1200, -600, 800]),
        ('Interest Paid', [-540, -540, -540, -540, -540]),
        ('Dividends Paid', [0, -1200, 0, 0, -1500]),
        ('Net Financing Cash Flow', [960, -2540, 660, -1140, -1240]),
        ('', []),
        ('Net Change in Cash', [2380, -760, 1715, 360, 905]),
        ('Opening Cash Balance', [15200, 17580, 16820, 18535, 18895]),
        ('Closing Cash Balance', [17580, 16820, 18535, 18895, 19800]),
    ]

    row = 5
    for item, values in cf_items:
        ws.cell(row, 1, item)

        if item in ['Operating Activities', 'Investing Activities', 'Financing Activities']:
            ws.cell(row, 1).font = Font(bold=True, size=12)
            ws.merge_cells(f'A{row}:F{row}')
        elif item in ['Net Operating Cash Flow', 'Net Investing Cash Flow', 'Net Financing Cash Flow',
                       'Net Change in Cash', 'Closing Cash Balance']:
            ws.cell(row, 1).font = Font(bold=True)
            for col in range(1, 7):
                ws.cell(row, col).border = Border(top=Side(style='medium'))

        for col, val in enumerate(values, 2):
            ws.cell(row, col, val)
            ws.cell(row, col).number_format = '#,##0'

            if val < 0 and item not in ['Closing Cash Balance', 'Opening Cash Balance']:
                ws.cell(row, col).font = Font(color="C00000")
            elif val > 0 and item in ['Net Operating Cash Flow', 'Closing Cash Balance']:
                ws.cell(row, col).font = Font(color="00B050", bold=True)

        row += 1

    # Liquidity Ratios
    ws[f'A{row+2}'] = "LIQUIDITY METRICS"
    ws[f'A{row+2}'].font = Font(size=14, bold=True)

    ws[f'A{row+4}'] = "Current Ratio"
    ws[f'B{row+4}'] = 2.4
    ws[f'B{row+4}'].number_format = '0.0'

    ws[f'A{row+5}'] = "Quick Ratio"
    ws[f'B{row+5}'] = 1.8
    ws[f'B{row+5}'].number_format = '0.0'

    ws[f'A{row+6}'] = "Cash Conversion Cycle (days)"
    ws[f'B{row+6}'] = 42

    ws[f'A{row+7}'] = "Days Sales Outstanding"
    ws[f'B{row+7}'] = 35

    ws[f'A{row+8}'] = "Days Payable Outstanding"
    ws[f'B{row+8}'] = 28

    # Auto-fit columns
    for col in range(1, 7):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22

    wb.save(f"{output_dir}/Cash_Flow_Analysis.xlsx")
    print("✓ Created Cash_Flow_Analysis.xlsx")

def create_regional_performance_matrix():
    """Create regional performance dashboard with geographic breakdown."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Regional Performance"

    # Title
    ws['A1'] = "REGIONAL PERFORMANCE MATRIX - GCC OPERATIONS"
    ws['A1'].font = Font(size=18, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws.merge_cells('A1:J1')
    ws.row_dimensions[1].height = 35

    # Headers
    headers = ['Region', 'Revenue (AED M)', 'YoY Growth %', 'Gross Margin %',
               'EBITDA Margin %', 'Market Share %', 'Customer Count',
               'Avg Transaction', 'Stores', 'Performance Rating']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(3, col, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Regional data
    regions = [
        ('UAE - Dubai', 45.2, 15.3, 42.5, 18.2, 23.4, 125000, 385, 18),
        ('UAE - Abu Dhabi', 28.7, 12.8, 40.1, 16.8, 18.9, 78000, 412, 12),
        ('KSA - Riyadh', 32.5, 18.5, 38.9, 15.4, 21.2, 142000, 298, 15),
        ('KSA - Jeddah', 18.3, 14.2, 39.7, 16.1, 19.5, 89000, 325, 9),
        ('Kuwait', 8.9, 9.8, 37.2, 14.8, 15.2, 34000, 358, 5),
        ('Qatar', 7.2, 11.5, 41.8, 17.5, 17.8, 28000, 395, 4),
        ('Bahrain', 4.8, 8.3, 36.5, 13.9, 14.1, 18000, 342, 3),
        ('Oman', 5.7, 10.2, 38.1, 15.2, 16.3, 22000, 368, 4),
    ]

    row = 4
    for region_data in regions:
        for col, val in enumerate(region_data, 1):
            ws.cell(row, col, val)

            # Format numbers
            if col == 2:  # Revenue
                ws.cell(row, col).number_format = '#,##0.0'
            elif col in [3, 4, 5, 6]:  # Percentages
                ws.cell(row, col).number_format = '0.0"%"'
            elif col in [7, 9]:  # Counts
                ws.cell(row, col).number_format = '#,##0'
            elif col == 8:  # Transaction value
                ws.cell(row, col).number_format = '#,##0'

            # Conditional formatting for growth
            if col == 3:
                if val >= 15:
                    ws.cell(row, col).fill = PatternFill(start_color="C6EFCE", fill_type="solid")
                    ws.cell(row, col).font = Font(color="006100", bold=True)
                elif val >= 10:
                    ws.cell(row, col).fill = PatternFill(start_color="FFEB9C", fill_type="solid")

        # Performance rating
        avg_score = (region_data[2] + region_data[3] + region_data[4]) / 3
        if avg_score >= 15:
            rating = "★★★★★"
            color = "00B050"
        elif avg_score >= 12:
            rating = "★★★★☆"
            color = "92D050"
        elif avg_score >= 10:
            rating = "★★★☆☆"
            color = "FFC000"
        else:
            rating = "★★☆☆☆"
            color = "FF6600"

        ws.cell(row, 10, rating)
        ws.cell(row, 10).font = Font(color=color, bold=True, size=14)

        row += 1

    # Totals
    ws.cell(row, 1, "TOTAL / AVERAGE")
    ws.cell(row, 1).font = Font(bold=True)
    ws.cell(row, 2, sum(r[1] for r in regions))
    ws.cell(row, 2).number_format = '#,##0.0'
    ws.cell(row, 2).font = Font(bold=True)

    for col in range(1, 11):
        ws.cell(row, col).border = Border(top=Side(style='thick'))

    # Auto-fit columns
    for col in range(1, 11):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 16

    ws.column_dimensions['A'].width = 20

    wb.save(f"{output_dir}/Regional_Performance_Matrix.xlsx")
    print("✓ Created Regional_Performance_Matrix.xlsx")

def create_yoy_variance_analysis():
    """Create comprehensive YoY variance analysis."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "YoY Variance"

    # Title
    ws['A1'] = "YEAR-OVER-YEAR VARIANCE ANALYSIS - DETAILED BREAKDOWN"
    ws['A1'].font = Font(size=18, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws.merge_cells('A1:H1')
    ws.row_dimensions[1].height = 35

    # Headers
    headers = ['Account', '2024 (AED M)', '2025 (AED M)', 'Variance (M)',
               'Variance %', 'Trend', 'Analysis', 'Action']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(3, col, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Variance data
    data = [
        ('Revenue', 113.5, 127.5, 'Favorable', 'Volume growth + price optimization'),
        ('COGS', 68.1, 72.3, 'Unfavorable', 'Raw material inflation 6.2%'),
        ('Gross Profit', 45.4, 55.2, 'Favorable', 'Improved product mix'),
        ('Sales & Marketing', 14.2, 15.8, 'Unfavorable', 'Digital marketing expansion'),
        ('General & Admin', 9.5, 10.2, 'Unfavorable', 'Headcount increase'),
        ('R&D', 2.8, 3.5, 'Unfavorable', 'Innovation pipeline investment'),
        ('EBITDA', 19.9, 25.7, 'Favorable', 'Operating leverage + efficiency'),
        ('Depreciation', 4.2, 4.2, 'Neutral', 'Stable CapEx base'),
        ('EBIT', 15.7, 21.5, 'Favorable', 'Strong operational performance'),
        ('Interest Expense', 2.2, 2.2, 'Neutral', 'Fixed rate debt structure'),
        ('Net Profit', 10.8, 15.5, 'Favorable', 'Margin expansion + growth'),
    ]

    row = 4
    for account, val_2024, val_2025, analysis_type, comment in data:
        ws.cell(row, 1, account)
        ws.cell(row, 1).font = Font(bold=True if account in ['Revenue', 'Gross Profit', 'EBITDA', 'EBIT', 'Net Profit'] else False)

        ws.cell(row, 2, val_2024)
        ws.cell(row, 2).number_format = '#,##0.0'

        ws.cell(row, 3, val_2025)
        ws.cell(row, 3).number_format = '#,##0.0'

        variance = val_2025 - val_2024
        ws.cell(row, 4, variance)
        ws.cell(row, 4).number_format = '#,##0.0'
        ws.cell(row, 4).font = Font(color="00B050" if variance > 0 else "C00000")

        var_pct = (variance / val_2024) * 100
        ws.cell(row, 5, var_pct)
        ws.cell(row, 5).number_format = '0.0"%"'
        ws.cell(row, 5).font = Font(color="00B050" if var_pct > 0 else "C00000", bold=True)

        # Trend arrow
        if analysis_type == 'Favorable':
            trend = "↑"
            color = "00B050"
        elif analysis_type == 'Unfavorable':
            trend = "↓"
            color = "C00000"
        else:
            trend = "→"
            color = "666666"

        ws.cell(row, 6, trend)
        ws.cell(row, 6).font = Font(size=16, color=color, bold=True)
        ws.cell(row, 6).alignment = Alignment(horizontal="center")

        ws.cell(row, 7, comment)

        action = "Monitor" if analysis_type == 'Favorable' else "Review" if analysis_type == 'Unfavorable' else "No action"
        ws.cell(row, 8, action)

        if account in ['Gross Profit', 'EBITDA', 'EBIT', 'Net Profit']:
            for col in range(1, 9):
                ws.cell(row, col).border = Border(bottom=Side(style='medium'))

        row += 1

    # Auto-fit columns
    for col in range(1, 9):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 16

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['G'].width = 35

    wb.save(f"{output_dir}/YoY_Variance_Analysis.xlsx")
    print("✓ Created YoY_Variance_Analysis.xlsx")

if __name__ == "__main__":
    print("Generating Finance Dashboards...")
    create_executive_summary_dashboard()
    create_cash_flow_analysis()
    create_regional_performance_matrix()
    create_yoy_variance_analysis()
    print(f"\n✅ All files created in: {output_dir}")
    print("\nNext: Open these files and take screenshots!")
