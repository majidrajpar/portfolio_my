#!/usr/bin/env python3
"""
Generate impressive Audit Tools Excel dashboards for portfolio screenshots.
Creates multiple worksheets with charts, conditional formatting, and professional styling.
"""
import os
from datetime import datetime, timedelta
import random
import string

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
except ImportError:
    print("Installing openpyxl...")
    os.system("pip install openpyxl")
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule

# Create output directory
output_dir = r"C:\Users\sorat\Desktop\Coding\portfolio_my\temp_excel_outputs"
os.makedirs(output_dir, exist_ok=True)

def create_branch_audit_checklist():
    """Create comprehensive branch audit checklist with scoring."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Branch Audit Checklist"

    # Title
    ws['A1'] = "BRANCH AUDIT CHECKLIST - COMPREHENSIVE CONTROL ASSESSMENT"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells('A1:H1')
    ws.row_dimensions[1].height = 30

    # Metadata
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A3'] = "Audit Period: Q1 2025 | Branches Audited: 15 | Overall Compliance: 87%"
    ws['A3'].font = Font(italic=True, color="666666")

    # Headers
    headers = ['Branch ID', 'Control Area', 'Control Description', 'Risk Level',
               'Test Result', 'Score (%)', 'Findings', 'Action Required']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style='thick'))

    # Sample data
    branches = [f"BR-{str(i).zfill(3)}" for i in range(1, 16)]
    control_areas = ['Cash Management', 'Inventory Control', 'Revenue Recognition',
                     'Access Controls', 'Segregation of Duties', 'Compliance Documentation']
    risk_levels = ['Critical', 'High', 'Medium', 'Low']
    test_results = ['Pass', 'Fail', 'Warning', 'N/A']

    row = 6
    for branch in branches:
        for _ in range(random.randint(3, 6)):  # 3-6 controls per branch
            ws.cell(row, 1, branch)
            ws.cell(row, 2, random.choice(control_areas))
            ws.cell(row, 3, f"Test procedure #{random.randint(100, 999)}: Control effectiveness validation")

            risk = random.choice(risk_levels)
            ws.cell(row, 4, risk)

            # Color code risk levels
            risk_cell = ws.cell(row, 4)
            if risk == 'Critical':
                risk_cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
                risk_cell.font = Font(color="FFFFFF", bold=True)
            elif risk == 'High':
                risk_cell.fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
                risk_cell.font = Font(color="FFFFFF")
            elif risk == 'Medium':
                risk_cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            else:
                risk_cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

            result = random.choice(test_results)
            ws.cell(row, 5, result)

            # Score based on result
            if result == 'Pass':
                score = random.randint(90, 100)
            elif result == 'Warning':
                score = random.randint(70, 89)
            elif result == 'Fail':
                score = random.randint(0, 69)
            else:
                score = None

            ws.cell(row, 6, score)

            # Conditional formatting for scores
            if score:
                score_cell = ws.cell(row, 6)
                score_cell.number_format = '0"%"'
                if score >= 90:
                    score_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    score_cell.font = Font(color="006100")
                elif score >= 70:
                    score_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    score_cell.font = Font(color="9C5700")
                else:
                    score_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    score_cell.font = Font(color="9C0006")

            findings = random.randint(0, 5)
            ws.cell(row, 7, findings if findings > 0 else "None")

            ws.cell(row, 8, "Follow-up required" if findings > 0 else "No action needed")

            row += 1

    # Auto-fit columns
    for col in range(1, 9):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    # Add summary chart
    ws2 = wb.create_sheet("Summary Dashboard")
    ws2['A1'] = "AUDIT PERFORMANCE DASHBOARD"
    ws2['A1'].font = Font(size=18, bold=True)
    ws2.merge_cells('A1:F1')

    # KPIs
    ws2['A3'] = "Overall Compliance Rate"
    ws2['B3'] = "87%"
    ws2['B3'].font = Font(size=24, bold=True, color="00B050")

    ws2['A5'] = "Total Controls Tested"
    ws2['B5'] = row - 6

    ws2['A6'] = "High Risk Findings"
    ws2['B6'] = random.randint(5, 15)

    wb.save(f"{output_dir}/Branch_Audit_Checklist.xlsx")
    print("✓ Created Branch_Audit_Checklist.xlsx")

def create_risk_assessment_matrix():
    """Create risk assessment heatmap with detailed scoring."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Risk Assessment Matrix"

    # Title
    ws['A1'] = "ENTERPRISE RISK ASSESSMENT MATRIX - COSO ERM FRAMEWORK"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws.merge_cells('A1:J1')
    ws.row_dimensions[1].height = 30

    # Headers
    headers = ['Risk ID', 'Risk Category', 'Risk Description', 'Likelihood (1-5)',
               'Impact (1-5)', 'Inherent Risk', 'Control Effectiveness', 'Residual Risk',
               'Risk Owner', 'Mitigation Status']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Risk categories
    categories = ['Operational', 'Financial', 'Compliance', 'Strategic', 'Cybersecurity', 'Reputation']
    risk_descriptions = [
        'Revenue leakage in cash handling processes',
        'Inventory shrinkage exceeding acceptable thresholds',
        'Non-compliance with GDPR data protection requirements',
        'Market share erosion due to competitor actions',
        'Ransomware attack on critical infrastructure',
        'Negative media coverage impacting brand value',
        'Supplier concentration risk in supply chain',
        'Foreign exchange volatility affecting margins',
        'Labor law violations in multi-jurisdiction operations',
        'Strategic initiative failure - digital transformation',
        'Data breach exposing customer PII',
        'Product quality issues damaging reputation'
    ]

    owners = ['CFO', 'COO', 'CLO', 'CEO', 'CISO', 'CMO']
    mitigation_status = ['Complete', 'In Progress', 'Planned', 'Not Started']

    row = 4
    for i in range(len(risk_descriptions)):
        ws.cell(row, 1, f"RSK-{str(i+1).zfill(3)}")
        ws.cell(row, 2, categories[i % len(categories)])
        ws.cell(row, 3, risk_descriptions[i])

        likelihood = random.randint(1, 5)
        impact = random.randint(1, 5)
        ws.cell(row, 4, likelihood)
        ws.cell(row, 5, impact)

        # Inherent Risk = Likelihood × Impact
        inherent = likelihood * impact
        ws.cell(row, 6, inherent)

        # Color code inherent risk
        inherent_cell = ws.cell(row, 6)
        if inherent >= 20:
            inherent_cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
            inherent_cell.font = Font(color="FFFFFF", bold=True)
        elif inherent >= 12:
            inherent_cell.fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
            inherent_cell.font = Font(color="FFFFFF")
        elif inherent >= 6:
            inherent_cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        else:
            inherent_cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

        control_eff = random.randint(50, 95)
        ws.cell(row, 7, control_eff)
        ws.cell(row, 7).number_format = '0"%"'

        # Residual Risk = Inherent × (1 - Control Effectiveness)
        residual = round(inherent * (1 - control_eff/100), 1)
        ws.cell(row, 8, residual)

        # Color code residual risk
        residual_cell = ws.cell(row, 8)
        if residual >= 10:
            residual_cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
            residual_cell.font = Font(color="FFFFFF", bold=True)
        elif residual >= 5:
            residual_cell.fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
            residual_cell.font = Font(color="FFFFFF")
        elif residual >= 2:
            residual_cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        else:
            residual_cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

        ws.cell(row, 9, random.choice(owners))
        ws.cell(row, 10, random.choice(mitigation_status))

        row += 1

    # Auto-fit columns
    for col in range(1, 11):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 16

    ws.column_dimensions['C'].width = 45

    wb.save(f"{output_dir}/Risk_Assessment_Matrix.xlsx")
    print("✓ Created Risk_Assessment_Matrix.xlsx")

def create_revenue_assurance_dashboard():
    """Create revenue assurance dashboard with trend analysis."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Revenue Assurance"

    # Title
    ws['A1'] = "REVENUE ASSURANCE & LEAKAGE DETECTION DASHBOARD"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws.merge_cells('A1:H1')
    ws.row_dimensions[1].height = 30

    # KPIs
    ws['A3'] = "Monthly Revenue Recovery"
    ws['B3'] = "AED 267,450"
    ws['B3'].font = Font(size=18, bold=True, color="00B050")

    ws['D3'] = "YTD Recovery"
    ws['E3'] = "AED 3.2M"
    ws['E3'].font = Font(size=18, bold=True, color="00B050")

    ws['G3'] = "Leakage Rate"
    ws['H3'] = "1.8%"
    ws['H3'].font = Font(size=18, bold=True, color="C00000")

    # Headers
    headers = ['Month', 'Expected Revenue', 'Actual Revenue', 'Variance',
               'Variance %', 'Recovered', 'Leakage Points', 'Action Status']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Monthly data
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    base_revenue = 15000000  # AED 15M

    for i, month in enumerate(months, 7):
        ws.cell(i, 1, month)

        expected = base_revenue + random.randint(-500000, 500000)
        actual = expected - random.randint(50000, 500000)
        variance = actual - expected
        variance_pct = (variance / expected) * 100
        recovered = abs(variance) * random.uniform(0.5, 0.9)
        leakage_points = random.randint(3, 12)

        ws.cell(i, 2, expected)
        ws.cell(i, 2).number_format = '#,##0'

        ws.cell(i, 3, actual)
        ws.cell(i, 3).number_format = '#,##0'

        ws.cell(i, 4, variance)
        ws.cell(i, 4).number_format = '#,##0'
        ws.cell(i, 4).font = Font(color="C00000" if variance < 0 else "00B050")

        ws.cell(i, 5, variance_pct)
        ws.cell(i, 5).number_format = '0.0"%"'
        ws.cell(i, 5).font = Font(color="C00000" if variance_pct < 0 else "00B050")

        ws.cell(i, 6, recovered)
        ws.cell(i, 6).number_format = '#,##0'

        ws.cell(i, 7, leakage_points)

        ws.cell(i, 8, "Resolved" if recovered / abs(variance) > 0.7 else "In Progress")

    # Auto-fit columns
    for col in range(1, 9):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 16

    wb.save(f"{output_dir}/Revenue_Assurance_Dashboard.xlsx")
    print("✓ Created Revenue_Assurance_Dashboard.xlsx")

def create_findings_tracker():
    """Create audit findings tracker with aging analysis."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Audit Findings"

    # Title
    ws['A1'] = "AUDIT FINDINGS TRACKER - REMEDIATION STATUS"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws.merge_cells('A1:J1')
    ws.row_dimensions[1].height = 30

    # Headers
    headers = ['Finding ID', 'Issue Date', 'Severity', 'Process Area', 'Finding Description',
               'Root Cause', 'Management Action', 'Due Date', 'Days Open', 'Status']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Sample findings
    severities = ['Critical', 'High', 'Medium', 'Low']
    process_areas = ['Procurement', 'Payroll', 'Revenue', 'Inventory', 'Fixed Assets', 'AP/AR']
    findings = [
        'Segregation of duties violation in payment approval',
        'Missing supporting documentation for expenses',
        'Unapproved vendor additions to master file',
        'Overtime calculations not validated by supervisor',
        'Inventory count discrepancies exceeding tolerance',
        'Fixed asset register not reconciled to GL',
        'Credit notes issued without proper authorization',
        'Bank reconciliation not reviewed timely',
        'Journal entries lacking adequate description',
        'Access rights not revoked for terminated employees'
    ]

    statuses = ['Open', 'In Progress', 'Pending Verification', 'Closed']

    row = 4
    for i in range(25):
        ws.cell(row, 1, f"FND-2025-{str(i+1).zfill(3)}")

        issue_date = datetime.now() - timedelta(days=random.randint(10, 180))
        ws.cell(row, 2, issue_date.strftime('%Y-%m-%d'))

        severity = random.choice(severities)
        ws.cell(row, 3, severity)

        # Color code severity
        sev_cell = ws.cell(row, 3)
        if severity == 'Critical':
            sev_cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
            sev_cell.font = Font(color="FFFFFF", bold=True)
        elif severity == 'High':
            sev_cell.fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
            sev_cell.font = Font(color="FFFFFF")
        elif severity == 'Medium':
            sev_cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        else:
            sev_cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

        ws.cell(row, 4, random.choice(process_areas))
        ws.cell(row, 5, random.choice(findings))
        ws.cell(row, 6, "Process design deficiency" if random.random() > 0.5 else "Control not operating effectively")
        ws.cell(row, 7, "Implement automated control" if random.random() > 0.5 else "Enhanced manual review process")

        due_date = issue_date + timedelta(days=random.randint(30, 90))
        ws.cell(row, 8, due_date.strftime('%Y-%m-%d'))

        days_open = (datetime.now() - issue_date).days
        ws.cell(row, 9, days_open)

        # Color code aging
        age_cell = ws.cell(row, 9)
        if days_open > 90:
            age_cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
            age_cell.font = Font(color="FFFFFF", bold=True)
        elif days_open > 60:
            age_cell.fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
            age_cell.font = Font(color="FFFFFF")
        elif days_open > 30:
            age_cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

        ws.cell(row, 10, random.choice(statuses))

        row += 1

    # Auto-fit columns
    for col in range(1, 11):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 14

    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 35
    ws.column_dimensions['G'].width = 35

    wb.save(f"{output_dir}/Audit_Findings_Tracker.xlsx")
    print("✓ Created Audit_Findings_Tracker.xlsx")

if __name__ == "__main__":
    print("Generating Audit Tools Dashboards...")
    create_branch_audit_checklist()
    create_risk_assessment_matrix()
    create_revenue_assurance_dashboard()
    create_findings_tracker()
    print(f"\n✅ All files created in: {output_dir}")
    print("\nNext: Open these files and take screenshots!")
